from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import User, Member, Loans, Amortization, BackupLog, RestoreLog
from auditlog.models import LogEntry
from .serializers import UserSerializer, LoanSerializer, AmortizationSerializer, AuditLogSerializer
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from .serializers import MemberSerializer
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.exceptions import TokenError, InvalidToken
from django.contrib.auth import authenticate #database for users
from django.http import Http404
from api.utils import generate_amortization_schedule
from rest_framework import mixins
from rest_framework import generics
from django.conf import settings
from auditlog.middleware import set_actor
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook
from io import BytesIO
import datetime
from django.http import HttpResponse
import os
from api.backup_restore import backup_view, restore_view  # import the views
from rest_framework.response import Response
from django.db.models import Q

#---USERS--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

@permission_classes([IsAuthenticated])
# Get all users
@api_view(['GET'])
def user_list(request):
    if request.method == 'GET':
        users = User.objects.all()
        serializer = UserSerializer(users, many=True)
        return Response(serializer.data)
@permission_classes([IsAuthenticated])
# Get a single user by ID
@api_view(['GET'])
def user_detail(request, pk):
    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = UserSerializer(user)
        return Response(serializer.data)
@permission_classes([IsAuthenticated])
# Create a new user
@api_view(['POST'])
def create_user(request):
    if request.method == 'POST':
        serializer = UserSerializer(data=request.data)
    if serializer.is_valid():
        with set_actor(request.user):
            serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

from auditlog.context import set_actor

@permission_classes([IsAuthenticated])
@api_view(['PUT', 'PATCH'])
def update_user(request, pk):
    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({'detail': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    
    serializer = UserSerializer(user, data=request.data, partial=True)
    if serializer.is_valid():
       
        with set_actor(request.user):
            serializer.save()

        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#---Search--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

@api_view(['GET'])
def search_users(request, query):
  
    if query.isdigit():
        users = User.objects.filter(
            Q(firstname__icontains=query) |
            Q(lastname__icontains=query) |
            Q(id=int(query))  # Convert to an integer for numeric ID search
        )
    else:
        users = User.objects.filter(
            Q(firstname__icontains=query) |
            Q(lastname__icontains=query)
        )
    
    serializer = UserSerializer(users, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def search_members(request, query):
    if query.isdigit():
        members = Member.objects.filter(
            Q(firstname__icontains=query) |
            Q(lastname__icontains=query) |
            Q(id=int(query))
        )
    else:
        members = Member.objects.filter(
            Q(firstname__icontains=query) |
            Q(lastname__icontains=query)
        )

    serializer = MemberSerializer(members, many=True)
    return Response(serializer.data)
#---LOGIN/AUTH/JWT--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

@api_view(['POST'])
@permission_classes([AllowAny])
def login_user(request):
    username = request.data.get('username')
    password = request.data.get('password')

    user = authenticate(request, username=username, password=password)

    if user is not None:
        set_actor(user)
        refresh = RefreshToken.for_user(user)
        access_token = str(refresh.access_token)

        response = Response({
            "message": "Login successful",
            "user": user.username,  # Add this line
        }, status=status.HTTP_200_OK)

        # Pull config from SIMPLE_JWT
        cookie_name = settings.SIMPLE_JWT.get("AUTH_COOKIE", "access_token")
        cookie_secure = settings.SIMPLE_JWT.get("AUTH_COOKIE_SECURE", False)
        cookie_httponly = settings.SIMPLE_JWT.get("AUTH_COOKIE_HTTP_ONLY", True)
        cookie_samesite = settings.SIMPLE_JWT.get("AUTH_COOKIE_SAMESITE", "Lax")
        access_token_lifetime = int(settings.SIMPLE_JWT["ACCESS_TOKEN_LIFETIME"].total_seconds())
        refresh_token_lifetime = int(settings.SIMPLE_JWT["REFRESH_TOKEN_LIFETIME"].total_seconds())

        # Set Access Token in Cookie
        response.set_cookie(
            key=cookie_name,
            value=access_token,
            httponly=cookie_httponly,
            secure=cookie_secure,
            samesite=cookie_samesite,
            max_age=access_token_lifetime,
            path='/'
        )

        # Optionally set Refresh Token in Cookie too
        response.set_cookie(
            key='refresh_token',
            value=str(refresh),
            httponly=cookie_httponly,
            secure=cookie_secure,
            samesite=cookie_samesite,
            max_age=refresh_token_lifetime,
            path='/'
        )

        return response
    else:
        return Response({"message": "Invalid credentials"}, status=status.HTTP_401_UNAUTHORIZED)
    
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    response = Response({'message': 'Logged out successfully'}, status=200)
    
    response.delete_cookie('access_token')
    response.delete_cookie('refresh_token')
    
    # Optional: also set expired manually (double safety)
    response.set_cookie('access_token', '', expires=0, httponly=True)
    response.set_cookie('refresh_token', '', expires=0, httponly=True)

    return response

@api_view(['POST'])
@permission_classes([AllowAny])
def refresh_token_view(request):
    refresh_token = request.COOKIES.get('refresh_token')
    if not refresh_token:
        return Response({'error': 'No refresh token'}, status=401)

    try:
        cookie_secure = settings.SIMPLE_JWT.get("AUTH_COOKIE_SECURE", False)
        cookie_httponly = settings.SIMPLE_JWT.get("AUTH_COOKIE_HTTP_ONLY", True)
        cookie_samesite = settings.SIMPLE_JWT.get("AUTH_COOKIE_SAMESITE", "Lax")
        access_token_lifetime = int(settings.SIMPLE_JWT["ACCESS_TOKEN_LIFETIME"].total_seconds())

        refresh = RefreshToken(refresh_token)
        access = str(refresh.access_token)

        res = Response({'access': access}, status=200)
        res.set_cookie(
            'access_token',
            access,
            httponly=cookie_httponly,
            samesite=cookie_samesite,
            secure=cookie_secure,
            max_age=access_token_lifetime, # 👈 Make it persist
            path="/"
        )

        return res
    except Exception:
        return Response({'detail': 'Invalid refresh token'}, status=403)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def protected_view(request):
    return Response({
        "authenticated": True,
        "user": request.user.username,
        'usertype': request.user.usertype
    })
    
#---MEMBERS--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# View to get all members (GET request)
@api_view(['GET'])
def get_members(request):
    members = Member.objects.all()  # Fetch all members from the database
    serializer = MemberSerializer(members, many=True)  # Serialize the members
    return Response(serializer.data)  # Return the serialized data as JSON

@api_view(['POST'])
def create_member(request):
    serializer = MemberSerializer(data=request.data)
    if serializer.is_valid():
        with set_actor(request.user):
            serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@permission_classes([IsAuthenticated])
# View to update an existing member
@api_view(['PUT', 'PATCH'])
def update_member(request, pk):

    try:
        member = Member.objects.get(pk=pk)  # Get the member by primary key (ID)
    except Member.DoesNotExist:
        return Response({'detail': 'Member not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = MemberSerializer(member, data=request.data, partial=True)  # Use the MemberSerializer
    if serializer.is_valid():
        # Use 'with' to set the actor context during save
        with set_actor(request.user):
            serializer.save()

        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#---LOANS--------------------------------------------------------------------------------------------------------------------------------------------------------------------------


# View to get all loans (GET request)
@api_view(['GET'])
def get_loan(request):

    loans = Loans.objects.all()  # Fetch all loans from the database
    serializer = LoanSerializer(loans, many=True)  # Serialize the loans
    return Response(serializer.data)  # Return the serialized data as JSON

@api_view(['GET'])
def search_loan(request, pk):
    try:
        loan = Loans.objects.get(pk=pk)  # Get the loan by primary key (ID)
    except Loans.DoesNotExist:
        return Response({'detail': 'Loan not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = LoanSerializer(loan)  # Serialize the loan
    return Response(serializer.data)  # Return the serialized data

@api_view(['POST'])
def create_loan(request):
    serializer = LoanSerializer(data=request.data)
    if serializer.is_valid():
        with set_actor(request.user):
            serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT', 'PATCH'])
def update_loan(request, pk):
    try:
        loan = Loans.objects.get(pk=pk)
    except Loans.DoesNotExist:
        return Response({'detail': 'Loan not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = LoanSerializer(loan, data=request.data, partial=True)
    if serializer.is_valid():
        with set_actor(request.user):
            serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def amortization_list(request, pk):
    try:
        amortizations = Amortization.objects.filter(loan_id=pk)
        serializer = AmortizationSerializer(amortizations, many=True)
        return Response(serializer.data)
    except Amortization.DoesNotExist:
        return Response({"detail": "No amortization records found for this loan."}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
def create_amortization_schedule(request, pk):
    loan = Loans.objects.get(pk=pk)
#Checks loan status
    if loan.status != 'released':
        return Response({"error": "Loan not released."}, status=400)

    if Amortization.objects.filter(loan=loan).exists():
        return Response({"error": "Schedule already exists."}, status=400)

    amortization = request.data.get('amortization')
    if amortization is None:
        return Response({"error": "Amortization value required."}, status=400)

    try:
        amortization = float(amortization)
    except ValueError:
        return Response({"error": "Invalid amortization value."}, status=400)

 
    generate_amortization_schedule(loan, amortization) #Calls this function to generate schedule

    # After generating, query the created amortizations
    amortizations = Amortization.objects.filter(loan=loan).order_by('seq')
    serializer = AmortizationSerializer(amortizations, many=True)
    return Response(serializer.data, status=201)

#---AUDIT LOGS--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

@permission_classes([IsAuthenticated])
@api_view(['GET'])
def get_audit_logs(request):
    logs = LogEntry.objects.all().order_by('-timestamp')  # Gets all logs, latest first
    serializer = AuditLogSerializer(logs, many=True)
    return Response(serializer.data)

#---REPORT GENERATION--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def replace_first_occurrence(ws, target_text, replacement, occurrence=1):
    """
    Searches the worksheet for cells whose value exactly equals target_text.
    When the nth occurrence (n = occurrence) is found, that cell’s value is replaced.
    """
    count = 0
    # Iterate over all cells in the worksheet row by row
    for row in ws.iter_rows():
        for cell in row:
            # Check if cell value matches the target text exactly (case sensitive, trimmed)
            if cell.value and str(cell.value).strip() == target_text:
                count += 1
                # When the target occurrence is reached, replace the cell value and return True
                if count == occurrence:
                    cell.value = replacement
                    return True
    # If the target occurrence is not found, return False
    return False


def set_merged_cell_value(ws, cell_coordinate, value):
    """
    Writes a value into the cell specified by cell_coordinate.
    If that cell belongs to a merged range, the value is written into the top-left cell of that range.
    """
    # Loop through all merged cell ranges in the worksheet
    for merged_range in ws.merged_cells.ranges:
        # Check if the target cell coordinate belongs to this merged range
        if cell_coordinate in merged_range:
            # Get the boundaries of the merged range (top-left corner)
            min_row, min_col, _, _ = merged_range.bounds
            # Set the value to the top-left cell of the merged range
            ws.cell(row=min_row, column=min_col, value=value)
            return
    # If the cell is not merged, set the value directly
    ws[cell_coordinate].value = value


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_loan_report(request, loan_id):
    """
    Generates a Loan Report using the pre-formatted "Report Format.xlsx" template.

    -- Top Section Mapping --
      • Replace the cell with literal "loan_id" with "Loan: {loan_id}".
      • Replace "date of download" with today's date (dd-mm‑yyyy).
      • (Do not do any placeholder replacement for "Start of Payment" or "Maturity Date".)
      • Replace the 2nd occurrence of "Interest" with the loan’s interest value.
      • Set cell P11 with the Loan Amount (for formula inputs) and R33 with the same value.
      • Set R16 with the term and O24 with the loan type; clear O26.
      • --- FINAL DATE REVISION ---
           Forcefully assign the **actual date objects** from the loan record to cells R27 and R28.
           (Here, R27 will get **payment_start_date** [the correct field name] and
            R28 will get **maturity_date**. Their number formats are forced to "DD-MM-YYYY".)

    -- Amortization Table Mapping (records start at row 34) --
      • Column A (index 1): record.seq
      • Column C (index 3): record.due_date (formatted as dd‑mm‑yyyy)
      • Column O (index 15): record.principal
      • Column P (index 16): record.interest
      • Column Q (index 17): record.amortization
      • Column R (index 18): record.remaining_balance

    All date cells are forced to display as "DD-MM-YYYY".
    """
    # Retrieve the loan instance by primary key or return 404 if not found
    loan = get_object_or_404(Loans, pk=loan_id)

    # Get today's date formatted as string dd-mm-yyyy for header replacement
    today = datetime.date.today().strftime("%d-%m-%Y")

    # Load the Excel template workbook
    template_path = os.path.join(settings.BASE_DIR, "template", "excel", "Report Format.xlsx")
    wb = load_workbook(template_path)
    ws = wb.active

    # Replace placeholders in the template's header cells
    replace_first_occurrence(ws, "loan_id", f"Loan: {loan.id}", occurrence=1)
    replace_first_occurrence(ws, "date of download", today, occurrence=1)
    replace_first_occurrence(ws, "Interest", loan.interest, occurrence=2)

    # Set specific cells with loan data; handle merged cells properly
    set_merged_cell_value(ws, "P11", loan.loan_amount)  # Loan amount for formulas
    ws["R33"].value = loan.loan_amount  # Loan amount repeated
    set_merged_cell_value(ws, "R16", loan.term)  # Loan term in months
    set_merged_cell_value(ws, "O24", loan.loan_type)  # Loan type
    set_merged_cell_value(ws, "O26", "")  # Clear this cell

    # Set actual date objects for payment start and maturity dates with formatting
    ws["R27"].value = loan.payment_start_date
    ws["R27"].number_format = "DD-MM-YYYY"

    ws["R28"].value = loan.maturity_date
    ws["R28"].number_format = "DD-MM-YYYY"

    # Fetch amortization records related to the loan, ordered by sequence number
    amortizations = Amortization.objects.filter(loan=loan).order_by("seq")

    current_row = 34  # Starting row for amortization table in the sheet
    for record in amortizations:
        # Write amortization schedule data to the correct columns and rows
        ws.cell(row=current_row, column=1, value=record.seq)  # Sequence number
        ws.cell(row=current_row, column=3, value=record.due_date)  # Due date
        ws.cell(row=current_row, column=3).number_format = "DD-MM-YYYY"  # Date format
        ws.cell(row=current_row, column=15, value=record.principal)  # Principal amount
        ws.cell(row=current_row, column=16, value=record.interest)  # Interest amount
        ws.cell(row=current_row, column=17, value=record.amortization)  # Amortization payment
        ws.cell(row=current_row, column=18, value=record.remaining_balance)  # Remaining balance
        current_row += 1  # Move to next row for next record

    # Save the workbook to an in-memory bytes buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Prepare HTTP response with the Excel file for download
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=Loan_Report_{loan.id}.xlsx'
    return response
#---BACKUP & RESTORE--------------------------------------------------------------------------------------------------------------------------------------------------------------------------

@api_view(["GET"])
def backup_view(request):
    try:
        backup_file_path = full_backup()  # Run backup, get file path
        return FileResponse(
            open(backup_file_path, "rb"),  # Open backup file
            content_type="application/gzip",  # Set content type for gzip
            as_attachment=True,  # Force download
            filename=os.path.basename(backup_file_path)  # Filename for download
        )
    except Exception as e:
        return JsonResponse({"detail": f"Backup failed: {e}"}, status=500)
    

@api_view(['GET'])
def last_backup_time(request):
    try:
        last_backup = BackupLog.objects.latest('backup_time')  # Get latest backup log
        return Response({"last_backup": last_backup.backup_time})
    except BackupLog.DoesNotExist:
        return Response({"last_backup": None})
    

@api_view(["GET"])
def last_restore_time(request):
    try:
        last_restore = RestoreLog.objects.latest('timestamp')  # Get latest restore log
        return Response({
            "last_restore_timestamp": last_restore.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        })
    except RestoreLog.DoesNotExist:
        return Response({"last_restore_timestamp": None})
